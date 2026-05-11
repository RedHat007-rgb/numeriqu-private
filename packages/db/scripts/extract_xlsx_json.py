import json
import sys
from datetime import date, datetime

from openpyxl import load_workbook


def _json_default(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    raise TypeError(f"Object of type {type(value).__name__} is not JSON serializable")


def sheet_to_records(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))

    # drop fully-empty trailing rows
    while rows and not any(v is not None for v in rows[-1]):
        rows.pop()

    if not rows:
        return {"headers": [], "records": []}

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    records = []
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        rec = {}
        for i, key in enumerate(headers):
            if not key:
                continue
            rec[key] = row[i] if i < len(row) else None
        records.append(rec)
    return {"headers": headers, "records": records}


def main():
    if len(sys.argv) < 2:
        print("Usage: extract_xlsx_json.py <file.xlsx>", file=sys.stderr)
        return 2

    path = sys.argv[1]
    wb = load_workbook(path, data_only=True, read_only=True)

    out = {"sheets": {}}
    for name in wb.sheetnames:
        ws = wb[name]
        out["sheets"][name] = sheet_to_records(ws)

    json.dump(out, sys.stdout, default=_json_default)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

